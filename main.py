import openpyxl
import networkx as nx
from numpy import exp, sqrt
import matplotlib.pyplot as plt

def min_weights(edge_list):
    # Функция поиска инцендентной линии с минимальной длиной
    # получает список линий инцендентных узлу. Возвращает длину линии с минмальным растоянием и название узла к
    # которому она идет
    edge_list = list(edge_list)
    weights = []
    for edge in edge_list:
        weights.append(edge[2]['weight'])
    return (min(weights), edge_list[weights.index(min(weights))][1])

def min_cyicle(node):
    # Функция поиска цикла с минимальной длиной (жадным алгоритмом) начиная с опредененного узла
    # на вход поступает номер узла node
    # на выходе получается перечень узлов через который проходит цикл - список cycle и его длина S
    cyicle = [node]
    S = 0
    G_1 = G.copy()
    node_list = list(G_1.nodes)
    #print(nx.to_edgelist(G_1, node))
    while len(node_list) > 1:
        new_hope = min_weights(nx.to_edgelist(G_1, node))
        S += new_hope[0]
        cyicle.append(str(new_hope[1]))
        node_list.remove(str(node))
        G_1.remove_node(node)
        node = str(new_hope[1])
    S += G.edges[str(cyicle[0]), str(node)]['weight']
    print(cyicle, S)
    return (cyicle, S)


# открываю файл с координатами узлов (колонки обозначают: первая название узла, вторая координату по X,
# третья координата по Y)
wb = openpyxl.load_workbook(filename='point.xlsx')
# открываю активный лист
ws = wb.active
# создаю пустой список
data = []
# определяю количество узлов по количеству заолненных строк
N_point = ws.max_row
# собираю значения всех ячеек в список data
for row in ws.values:
   for value in row:
       data.append(value)

# создаю словарь для координат узлов
Nodes = {}
for i in range(0, len(data), 3):
    # заполняю словарь: первое значение ключ - номер узла (с нуля) втрое и третье заначение координаты
    Nodes.setdefault(str(data[i]), (data[i+1], data[i+2]))

# полносвязный граф на основе словаря
G = nx.complete_graph(Nodes)

# создаю список линий
Lines = []
for edge in G.edges:
    distance = sqrt((Nodes[edge[0]][0] - Nodes[edge[1]][0]) ** 2 + (Nodes[edge[0]][1] - Nodes[edge[1]][1]) ** 2)
    # расчет длин ребер между узлами
    line = (edge[0], edge[1], distance)
    Lines.append(line)
G.add_weighted_edges_from(Lines)
# добавляем длины линий в граф

path_list = []
# создаем список минимальных циклов
distance_list = []
# создаем список длин циклов полученных циклов
for node in G.nodes:
    new_cyicle = min_cyicle(node)
    # находим жадным алгоритмом минимальный цикл для выбранного узла node
    path_list.append(new_cyicle[0])
    # заполняем список минимальных циклов
    distance_list.append(new_cyicle[1])
    # заполняем список длин циклов

# присваиваем полученный результат
res_cyicle = path_list[distance_list.index(min(distance_list))]
res_distance = min(distance_list)

# выводим результат в терминал
print('\n', 'Минимальный цикл ', res_cyicle,'\n'
      ,'Минимальная длина цикла =', res_distance, '\n')

# сохраняем результат в файл exel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'результат'
ws['A1'] = 'Минимальная длина цикла'
ws['A2'] = res_distance
ws['A3'] = 'Минимальный цикл'
res_cyicle_str = ''
for i in range(0, len(res_cyicle)):
   res_cyicle_str += str(res_cyicle[i])
ws['A4'] = res_cyicle_str
wb.save(filename='Results.xlsx')

# переделываю список линий цикла для отрисовки результата
res_cyicle_edges = []
for i in range(0, len(res_cyicle)):
    if i == 0:
        res_cyicle_edge = (res_cyicle[len(res_cyicle)-1], res_cyicle[0])
    else:
        res_cyicle_edge = (res_cyicle[i-1], res_cyicle[i])
    res_cyicle_edges.append(res_cyicle_edge)

# отрисовываем исходные данные и полученный результат
plt.subplot(121)
plt.title('Исходные данные: \n Количество узлов - %i.' %N_point) # печатаю заголовок рисунка
G_2 = G.copy()
G_2.clear_edges()
nx.draw(G_2, pos=Nodes, with_labels=True) # рисую граф с названиями и позициями узлов
plt.draw() # формирую рисунок
plt.subplot(122)
plt.title('Результат расчета: \n Минимальная длина цикла =  %i.' %res_distance) # печатаю заголовок рисунка
G_2.add_edges_from(res_cyicle_edges)
nx.draw(G_2, pos=Nodes, with_labels=True) # рисую граф с названиями и позициями узлов и минмальным циклом
plt.draw() # формирую рисунок
plt.show() # вывожу рисунок

